import os
import json
import re

# 此功能需要 openpyxl 库，请通过命令 "pip install openpyxl" 来安装
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                             QTreeWidget, QTreeWidgetItem, QAbstractItemView, QHeaderView,
                             QGroupBox, QInputDialog, QMessageBox, QMenu, QFileDialog, QLineEdit,
                             QStyledItemDelegate)
from PyQt6.QtGui import QIcon, QAction
from PyQt6.QtCore import Qt, pyqtSignal, QSize

REWARDS_FILE = "qianqian_rewards.json"


class LargeEditorDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QLineEdit):
            editor.setStyleSheet("font-size: 14px; padding: 5px;")
            editor.setMinimumHeight(30)
        return editor

    def sizeHint(self, option, index):
        size = super().sizeHint(option, index)
        size.setHeight(35)
        return size


class RewardsTab(QWidget):
    rewards_updated = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.daily_plan_tree = None
        self.current_plan_tree = None
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        # --- Top Button Bar ---
        top_button_layout = QHBoxLayout()
        add_root_btn = QPushButton("➕ 添加主计划")
        add_child_btn = QPushButton("➕ 添加子计划")
        remove_btn = QPushButton("➖ 删除选中计划")
        import_btn = QPushButton("📥 导入")
        export_btn = QPushButton("📤 导出")

        top_button_layout.addWidget(add_root_btn)
        top_button_layout.addWidget(add_child_btn)
        top_button_layout.addWidget(remove_btn)
        top_button_layout.addStretch()
        top_button_layout.addWidget(import_btn)
        top_button_layout.addWidget(export_btn)
        main_layout.addLayout(top_button_layout)

        # --- Tree Views ---
        plans_layout = QHBoxLayout()
        self.daily_plan_tree = self.create_plan_tree()
        daily_group = QGroupBox("📅 当天计划")
        daily_layout = QVBoxLayout(daily_group)
        daily_layout.addWidget(self.daily_plan_tree)

        self.current_plan_tree = self.create_plan_tree()
        current_group = QGroupBox("🎯 当前计划")
        current_layout = QVBoxLayout(current_group)
        current_layout.addWidget(self.current_plan_tree)

        plans_layout.addWidget(daily_group)
        plans_layout.addWidget(current_group)
        main_layout.addLayout(plans_layout)

        self.setLayout(main_layout)

        # Connect signals
        add_root_btn.clicked.connect(self.add_root_item)
        add_child_btn.clicked.connect(self.add_child_item)
        remove_btn.clicked.connect(self.remove_selected_item)
        import_btn.clicked.connect(self.import_data)
        export_btn.clicked.connect(self.export_data)

    def create_plan_tree(self):
        tree = QTreeWidget()
        tree.setHeaderLabels(["计划内容", "对应奖励"])
        tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        tree.itemChanged.connect(self.handle_item_changed)
        tree.customContextMenuRequested.connect(lambda pos, t=tree: self.show_context_menu(pos, t))

        delegate = LargeEditorDelegate(tree)
        tree.setItemDelegate(delegate)
        return tree

    def get_active_tree(self):
        if self.daily_plan_tree.hasFocus():
            return self.daily_plan_tree, "当天计划"
        if self.current_plan_tree.hasFocus():
            return self.current_plan_tree, "当前计划"

        # 如果都没有焦点，则弹窗让用户选择
        items = ["当天计划", "当前计划"]
        item, ok = QInputDialog.getItem(self, "选择列表", "请选择要操作的计划列表:", items, 0, False)
        if ok and item:
            if item == "当天计划":
                return self.daily_plan_tree, "当天计划"
            else:
                return self.current_plan_tree, "当前计划"
        return None, None

    def add_root_item(self):
        tree, _ = self.get_active_tree()
        if tree:
            self.add_item(tree)

    def add_child_item(self, parent_item=None):
        tree, _ = self.get_active_tree()
        if not tree: return

        if not parent_item:
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(self, "提示", "请先选择一个父计划。")
                return
            parent_item = selected[0]

        self.add_item(tree, as_child=True, parent_override=parent_item)

    def remove_selected_item(self, item_to_remove=None):
        tree, _ = self.get_active_tree()
        if not tree: return

        if not item_to_remove:
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(self, "提示", "请先选择要删除的计划。")
                return
            item_to_remove = selected[0]

        reply = QMessageBox.question(self, "确认删除", "确定要删除选中的计划及其所有子计划吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            (item_to_remove.parent() or tree.invisibleRootItem()).removeChild(item_to_remove)
            self.save_data()

    def add_item(self, tree, as_child=False, parent_override=None):
        parent = tree.invisibleRootItem()
        if as_child:
            if parent_override:
                parent = parent_override
            else:
                selected = tree.selectedItems()
                if not selected:
                    QMessageBox.information(self, "提示", "请先选择一个父计划。")
                    return
                parent = selected[0]

        plan_dialog = QInputDialog(self)
        plan_dialog.setWindowTitle("添加计划")
        plan_dialog.setLabelText("计划内容:")
        plan_dialog.resize(400, 100)
        ok1 = plan_dialog.exec()
        plan = plan_dialog.textValue()

        if ok1 and plan:
            reward_dialog = QInputDialog(self)
            reward_dialog.setWindowTitle("添加奖励")
            reward_dialog.setLabelText("对应奖励:")
            reward_dialog.resize(400, 100)
            ok2 = reward_dialog.exec()
            reward = reward_dialog.textValue()

            if ok2 is not None:
                new_item = QTreeWidgetItem(parent, [plan, reward])
                new_item.setFlags(new_item.flags() | Qt.ItemFlag.ItemIsEditable)
                if as_child:
                    parent.setExpanded(True)
                self.save_data()

    def handle_item_changed(self, item, column):
        self.save_data()

    def show_context_menu(self, position, tree):
        menu = QMenu()
        selected_item = tree.itemAt(position)

        if selected_item:
            add_child_action = QAction("➕ 在此项下添加子计划", self)
            add_child_action.triggered.connect(lambda: self.add_child_item(parent_item=selected_item))
            menu.addAction(add_child_action)

            delete_action = QAction("➖ 删除此计划", self)
            delete_action.triggered.connect(lambda: self.remove_selected_item(item_to_remove=selected_item))
            menu.addAction(delete_action)
            menu.addSeparator()

        add_root_action = QAction("➕ 添加主计划", self)
        add_root_action.triggered.connect(lambda: self.add_item(tree))
        menu.addAction(add_root_action)

        clear_all_action = QAction("🗑️ 清空所有计划", self)
        clear_all_action.triggered.connect(lambda: self.clear_tree(tree))
        menu.addAction(clear_all_action)

        menu.addSeparator()

        expand_all_action = QAction("📂 展开所有", self)
        expand_all_action.triggered.connect(lambda: self.expand_all_items(tree))
        menu.addAction(expand_all_action)

        collapse_all_action = QAction("📁 折叠所有", self)
        collapse_all_action.triggered.connect(lambda: self.collapse_all_items(tree))
        menu.addAction(collapse_all_action)

        menu.exec(tree.viewport().mapToGlobal(position))

    def clear_tree(self, tree):
        tree_name = "当天计划" if tree is self.daily_plan_tree else "当前计划"
        reply = QMessageBox.question(self, "确认清空", f"确定要清空【{tree_name}】中的所有计划吗？此操作不可撤销。",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            tree.clear()
            self.save_data()

    def expand_all_items(self, tree):
        tree.expandAll()

    def collapse_all_items(self, tree):
        tree.collapseAll()

    def get_data(self):
        """以字典形式返回树中的当前数据。"""
        return {
            "daily": self.tree_to_dict(self.daily_plan_tree),
            "current": self.tree_to_dict(self.current_plan_tree)
        }

    def load_data(self):
        data = {}
        if not os.path.exists(REWARDS_FILE):
            self.rewards_updated.emit(data)
            return
        try:
            with open(REWARDS_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.dict_to_tree(data.get("daily", []), self.daily_plan_tree)
            self.dict_to_tree(data.get("current", []), self.current_plan_tree)
        except (json.JSONDecodeError, FileNotFoundError):
            data = {}
            pass
        self.rewards_updated.emit(data)

    def save_data(self):
        data = self.get_data()
        with open(REWARDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        self.rewards_updated.emit(data)

    def tree_to_dict(self, tree):
        def recurse(parent_item):
            return [{
                "plan": parent_item.child(i).text(0),
                "reward": parent_item.child(i).text(1),
                "children": recurse(parent_item.child(i))
            } for i in range(parent_item.childCount())]

        if tree is None: return []
        return recurse(tree.invisibleRootItem())

    def dict_to_tree(self, data, tree):
        def recurse(parent_item, children_data):
            for item_data in children_data:
                child_item = QTreeWidgetItem(parent_item, [item_data["plan"], item_data["reward"]])
                child_item.setFlags(child_item.flags() | Qt.ItemFlag.ItemIsEditable)
                recurse(child_item, item_data.get("children", []))

        if tree is None: return
        tree.clear()
        recurse(tree.invisibleRootItem(), data)

    def import_data(self):
        items = ["当天计划", "当前计划"]
        item, ok = QInputDialog.getItem(self, "选择导入", "您想将计划导入到哪个列表？", items, 0, False)
        if not ok or not item:
            return

        tree = self.daily_plan_tree if item == "当天计划" else self.current_plan_tree
        tree_name = item

        file_path, _ = QFileDialog.getOpenFileName(self, f"从文件导入到 {tree_name}", "",
                                                   "所有支持的文件 (*.xlsx *.txt);;Excel 文件 (*.xlsx);;文本文件 (*.txt)")
        if not file_path:
            return

        if file_path.lower().endswith('.xlsx'):
            self.import_from_excel(file_path, tree, tree_name)
        elif file_path.lower().endswith('.txt'):
            self.import_from_txt(file_path, tree, tree_name)

    def import_from_txt(self, file_path, tree, tree_name):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()

            parent_item = tree.invisibleRootItem()
            imported_count = 0
            time_pattern = re.compile(r"^\s*\d{2}:\d{2}\s*-\s*\d{2}:\d{2}\s+")

            for line in lines:
                match = time_pattern.match(line)
                if match:
                    plan_text = line[match.end():].strip()
                    if plan_text:
                        new_item = QTreeWidgetItem(parent_item, [plan_text, ""])
                        new_item.setFlags(new_item.flags() | Qt.ItemFlag.ItemIsEditable)
                        imported_count += 1

            if imported_count > 0:
                self.save_data()
                QMessageBox.information(self, "成功", f"成功从文本文件合并导入 {imported_count} 条计划到 {tree_name}。")
            else:
                QMessageBox.warning(self, "导入提示", "在文件中没有找到符合格式 '时间段 内容' 的有效数据行。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"从 TXT 文件导入过程中发生错误: {e}")

    def import_from_excel(self, file_path, tree, tree_name):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导入功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            if sheet["A1"].value != "计划内容" or sheet["B1"].value != "对应奖励字":
                QMessageBox.warning(self, "格式错误",
                                    "导入的Excel文件格式不正确。\n第一行应为 '计划内容' 和 '对应奖励字'。")
                return

            # 不再清空 tree.clear()
            parent_stack = [tree.invisibleRootItem()]

            for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                plan_text, reward_text = (str(row[0]) if row[0] is not None else ""), (
                    str(row[1]) if row[1] is not None else "")
                if not plan_text: continue

                leading_spaces = len(plan_text) - len(plan_text.lstrip(' '))
                level = leading_spaces // 2
                clean_plan_text = plan_text.lstrip(' ')

                while level < len(parent_stack) - 1:
                    parent_stack.pop()

                if level > len(parent_stack) - 1: level = len(parent_stack) - 1

                parent = parent_stack[level]
                new_item = QTreeWidgetItem(parent, [clean_plan_text, reward_text])
                new_item.setFlags(new_item.flags() | Qt.ItemFlag.ItemIsEditable)

                if len(parent_stack) > level + 1:
                    parent_stack[level + 1] = new_item
                else:
                    parent_stack.append(new_item)

            self.save_data()
            QMessageBox.information(self, "成功", f"数据已成功合并导入到 {tree_name}。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入过程中发生错误: {e}")

    def export_data(self):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导出功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return

        items = ["当天计划", "当前计划"]
        item, ok = QInputDialog.getItem(self, "选择导出", "您想从哪个列表导出计划？", items, 0, False)
        if not ok or not item:
            return

        tree = self.daily_plan_tree if item == "当天计划" else self.current_plan_tree
        default_name = item

        file_path, _ = QFileDialog.getSaveFileName(self, f"导出 {default_name} 到 Excel", f"{default_name}.xlsx",
                                                   "Excel 文件 (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = default_name

            sheet["A1"] = "计划内容"
            sheet["B1"] = "对应奖励字"
            sheet.column_dimensions[get_column_letter(1)].width = 50
            sheet.column_dimensions[get_column_letter(2)].width = 30

            def write_items_recursive(parent_item, current_row, level=0):
                for i in range(parent_item.childCount()):
                    item = parent_item.child(i)
                    indent = "  " * level
                    plan_text = f"{indent}{item.text(0)}"
                    reward_text = item.text(1)
                    sheet[f"A{current_row}"] = plan_text
                    sheet[f"B{current_row}"] = reward_text
                    current_row += 1
                    current_row = write_items_recursive(item, current_row, level + 1)
                return current_row

            write_items_recursive(tree.invisibleRootItem(), 2)
            workbook.save(file_path)
            QMessageBox.information(self, "成功", f"{default_name} 已成功导出到:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误: {e}")

