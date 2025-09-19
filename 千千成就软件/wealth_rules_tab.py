import os
import json
import re

# 此功能需要 openpyxl 库，请通过命令 "pip install openpyxl" 来安装
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
                             QTreeWidget, QTreeWidgetItem, QAbstractItemView, QHeaderView,
                             QGroupBox, QInputDialog, QMessageBox, QMenu, QFileDialog, QLineEdit,
                             QStyledItemDelegate)
from PyQt6.QtGui import QIcon, QAction, QIntValidator
from PyQt6.QtCore import Qt, pyqtSignal, QSize

RULES_FILE = "rules.json"


class LargeEditorDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QLineEdit):
            editor.setStyleSheet("font-size: 14px; padding: 5px;")
            editor.setMinimumHeight(30)
            # 为第三列（索引为2）添加整数验证器
            if index.column() == 2:
                editor.setValidator(QIntValidator(parent))
        return editor

    def sizeHint(self, option, index):
        size = super().sizeHint(option, index)
        size.setHeight(35)
        return size


class WealthRulesTab(QWidget):
    rules_updated = pyqtSignal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.normal_rules_tree = None
        self.special_rules_tree = None
        self.term = "财富"
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout(self)

        # --- Top Button Bar ---
        top_button_layout = QHBoxLayout()
        add_root_btn = QPushButton("➕ 添加主规则")
        add_child_btn = QPushButton("➕ 添加子规则")
        remove_btn = QPushButton("➖ 删除选中规则")
        import_btn = QPushButton("📥 导入")
        export_btn = QPushButton("📤 导出")

        top_button_layout.addWidget(add_root_btn)
        top_button_layout.addWidget(add_child_btn)
        top_button_layout.addWidget(remove_btn)
        top_button_layout.addStretch()
        top_button_layout.addWidget(import_btn)
        top_button_layout.addWidget(export_btn)
        main_layout.addLayout(top_button_layout)

        # --- Tree Views ---
        plans_layout = QHBoxLayout()
        self.normal_rules_tree = self.create_plan_tree()
        self.normal_group = QGroupBox(f"📅 普通({self.term})奖励")
        daily_layout = QVBoxLayout(self.normal_group)
        daily_layout.addWidget(self.normal_rules_tree)

        self.special_rules_tree = self.create_plan_tree()
        self.special_group = QGroupBox(f"🎯 特殊({self.term})奖励")
        current_layout = QVBoxLayout(self.special_group)
        current_layout.addWidget(self.special_rules_tree)

        plans_layout.addWidget(self.normal_group)
        plans_layout.addWidget(self.special_group)
        main_layout.addLayout(plans_layout)

        self.setLayout(main_layout)

        # Connect signals
        add_root_btn.clicked.connect(self.add_root_item)
        add_child_btn.clicked.connect(self.add_child_item)
        remove_btn.clicked.connect(self.remove_selected_item)
        import_btn.clicked.connect(self.import_data)
        export_btn.clicked.connect(self.export_data)

    def apply_settings(self, style_config):
        self.term = style_config.get("term_display_mode", "财富")
        self.normal_group.setTitle(f"📅 普通({self.term})奖励")
        self.special_group.setTitle(f"🎯 特殊({self.term})奖励")
        headers = ["规则内容", "对应奖励", f"对应{self.term}"]
        self.normal_rules_tree.setHeaderLabels(headers)
        self.special_rules_tree.setHeaderLabels(headers)

    def create_plan_tree(self):
        tree = QTreeWidget()
        tree.setHeaderLabels(["规则内容", "对应奖励", f"对应{self.term}"])
        header = tree.header()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)

        tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        tree.itemChanged.connect(self.handle_item_changed)
        tree.customContextMenuRequested.connect(lambda pos, t=tree: self.show_context_menu(pos, t))

        delegate = LargeEditorDelegate(tree)
        tree.setItemDelegate(delegate)
        return tree

    def get_active_tree(self):
        if self.normal_rules_tree.hasFocus():
            return self.normal_rules_tree, f"普通({self.term})奖励"
        if self.special_rules_tree.hasFocus():
            return self.special_rules_tree, f"特殊({self.term})奖励"

        items = [f"普通({self.term})奖励", f"特殊({self.term})奖励"]
        item, ok = QInputDialog.getItem(self, "选择列表", "请选择要操作的规则列表:", items, 0, False)
        if ok and item:
            if item == f"普通({self.term})奖励":
                return self.normal_rules_tree, f"普通({self.term})奖励"
            else:
                return self.special_rules_tree, f"特殊({self.term})奖励"
        return None, None

    def add_root_item(self):
        tree, _ = self.get_active_tree()
        if tree:
            self.add_item(tree)

    def add_child_item(self, parent_item=None):
        tree, _ = self.get_active_tree()
        if not tree: return

        if not parent_item:
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(self, "提示", "请先选择一个父规则。")
                return
            parent_item = selected[0]

        self.add_item(tree, as_child=True, parent_override=parent_item)

    def remove_selected_item(self, item_to_remove=None):
        tree, _ = self.get_active_tree()
        if not tree: return

        if not item_to_remove:
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(self, "提示", "请先选择要删除的规则。")
                return
            item_to_remove = selected[0]

        reply = QMessageBox.question(self, "确认删除", "确定要删除选中的规则及其所有子规则吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            (item_to_remove.parent() or tree.invisibleRootItem()).removeChild(item_to_remove)
            self.save_data()

    def add_item(self, tree, as_child=False, parent_override=None):
        parent = tree.invisibleRootItem()
        if as_child:
            if parent_override:
                parent = parent_override
            else:
                selected = tree.selectedItems()
                if not selected:
                    QMessageBox.information(self, "提示", "请先选择一个父规则。")
                    return
                parent = selected[0]

        plan, ok1 = QInputDialog.getText(self, "添加规则", "规则内容:", QLineEdit.EchoMode.Normal, "")
        if not ok1 or not plan:
            return

        reward_text, ok2 = QInputDialog.getText(self, "添加奖励文字", "对应奖励:", QLineEdit.EchoMode.Normal, "")
        if not ok2:
            return

        reward_value, ok3 = QInputDialog.getInt(self, f"添加{self.term}", f"对应{self.term} (整数):", 0)
        if not ok3:
            return

        new_item = QTreeWidgetItem(parent, [plan, reward_text, str(reward_value)])
        new_item.setFlags(new_item.flags() | Qt.ItemFlag.ItemIsEditable)
        if as_child:
            parent.setExpanded(True)
        self.save_data()

    def handle_item_changed(self, item, column):
        self.save_data()

    def show_context_menu(self, position, tree):
        menu = QMenu()
        selected_item = tree.itemAt(position)

        if selected_item:
            add_child_action = QAction("➕ 在此项下添加子规则", self)
            add_child_action.triggered.connect(lambda: self.add_child_item(parent_item=selected_item))
            menu.addAction(add_child_action)

            delete_action = QAction("➖ 删除此规则", self)
            delete_action.triggered.connect(lambda: self.remove_selected_item(item_to_remove=selected_item))
            menu.addAction(delete_action)
            menu.addSeparator()

        add_root_action = QAction("➕ 添加主规则", self)
        add_root_action.triggered.connect(lambda: self.add_item(tree))
        menu.addAction(add_root_action)

        clear_all_action = QAction("🗑️ 清空所有规则", self)
        clear_all_action.triggered.connect(lambda: self.clear_tree(tree))
        menu.addAction(clear_all_action)

        menu.addSeparator()

        expand_all_action = QAction("📂 展开所有", self)
        expand_all_action.triggered.connect(lambda: self.expand_all_items(tree))
        menu.addAction(expand_all_action)

        collapse_all_action = QAction("📁 折叠所有", self)
        collapse_all_action.triggered.connect(lambda: self.collapse_all_items(tree))
        menu.addAction(collapse_all_action)

        menu.exec(tree.viewport().mapToGlobal(position))

    def clear_tree(self, tree):
        tree_name = f"普通({self.term})奖励" if tree is self.normal_rules_tree else f"特殊({self.term})奖励"
        reply = QMessageBox.question(self, "确认清空", f"确定要清空【{tree_name}】中的所有规则吗？此操作不可撤销。",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            tree.clear()
            self.save_data()

    def expand_all_items(self, tree):
        tree.expandAll()

    def collapse_all_items(self, tree):
        tree.collapseAll()

    def get_data(self):
        return {
            "normal": self.tree_to_dict(self.normal_rules_tree),
            "special": self.tree_to_dict(self.special_rules_tree)
        }

    def load_data(self):
        if not os.path.exists(RULES_FILE):
            # If the file does not exist, create and load default data
            default_data = {
                "normal": [
                    {"plan": "读书1个小时", "reward_text": "", "reward": "10", "children": []}
                ],
                "special": [
                    {"plan": "考过四/六级", "reward_text": "奖励新手机", "reward": "300", "children": []}
                ]
            }
            self.dict_to_tree(default_data.get("normal", []), self.normal_rules_tree)
            self.dict_to_tree(default_data.get("special", []), self.special_rules_tree)
            self.save_data()  # This will save the file and emit the signal
            return

        # If the file exists, load it
        data = {}
        try:
            with open(RULES_FILE, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.dict_to_tree(data.get("normal", []), self.normal_rules_tree)
            self.dict_to_tree(data.get("special", []), self.special_rules_tree)
        except (json.JSONDecodeError, FileNotFoundError):
            # In case of corruption or other errors, data remains {}
            pass

        self.rules_updated.emit(data)

    def save_data(self):
        data = self.get_data()
        with open(RULES_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        self.rules_updated.emit(data)

    def tree_to_dict(self, tree):
        def recurse(parent_item):
            return [{
                "plan": parent_item.child(i).text(0),
                "reward_text": parent_item.child(i).text(1),
                "reward": parent_item.child(i).text(2),
                "children": recurse(parent_item.child(i))
            } for i in range(parent_item.childCount())]

        if tree is None: return []
        return recurse(tree.invisibleRootItem())

    def dict_to_tree(self, data, tree):
        def recurse(parent_item, children_data):
            for item_data in children_data:
                child_item = QTreeWidgetItem(parent_item, [
                    item_data.get("plan", ""),
                    item_data.get("reward_text", ""),
                    str(item_data.get("reward", "0"))
                ])
                child_item.setFlags(child_item.flags() | Qt.ItemFlag.ItemIsEditable)
                recurse(child_item, item_data.get("children", []))

        if tree is None: return
        tree.clear()
        recurse(tree.invisibleRootItem(), data)

    def import_data(self):
        items = [f"普通({self.term})奖励", f"特殊({self.term})奖励"]
        item, ok = QInputDialog.getItem(self, "选择导入", "您想将规则导入到哪个列表？", items, 0, False)
        if not ok or not item:
            return

        tree = self.normal_rules_tree if item == f"普通({self.term})奖励" else self.special_rules_tree
        tree_name = item

        file_path, _ = QFileDialog.getOpenFileName(self, f"从文件导入到 {tree_name}", "",
                                                   "Excel 文件 (*.xlsx);;所有文件 (*)")
        if not file_path:
            return

        if file_path.lower().endswith('.xlsx'):
            self.import_from_excel(file_path, tree, tree_name)
        else:
            QMessageBox.warning(self, "文件类型不支持", "目前仅支持从 .xlsx 文件导入。")

    def import_from_excel(self, file_path, tree, tree_name):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导入功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            expected_header1 = "规则内容"
            expected_header2 = "对应奖励"
            expected_header3 = f"对应{self.term}"

            if (sheet["A1"].value != expected_header1 or
                    sheet["B1"].value != expected_header2 or
                    sheet["C1"].value != expected_header3):
                QMessageBox.warning(self, "格式错误",
                                    f"导入的Excel文件格式不正确。\n第一行应为 '{expected_header1}', '{expected_header2}' 和 '{expected_header3}'。")
                return

            parent_stack = [tree.invisibleRootItem()]
            imported_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 3: continue

                plan_text, reward_text, reward_val = (
                    str(row[0]) if row[0] is not None else "",
                    str(row[1]) if row[1] is not None else "",
                    row[2]
                )
                if not plan_text: continue

                try:
                    reward_int = int(reward_val) if reward_val is not None else 0
                except (ValueError, TypeError):
                    continue

                leading_spaces = len(plan_text) - len(plan_text.lstrip(' '))
                level = leading_spaces // 2
                clean_plan_text = plan_text.lstrip(' ')

                while level < len(parent_stack) - 1:
                    parent_stack.pop()
                if level > len(parent_stack) - 1:
                    level = len(parent_stack) - 1

                parent = parent_stack[level]
                new_item = QTreeWidgetItem(parent, [clean_plan_text, reward_text, str(reward_int)])
                new_item.setFlags(new_item.flags() | Qt.ItemFlag.ItemIsEditable)

                if len(parent_stack) > level + 1:
                    parent_stack[level + 1] = new_item
                else:
                    parent_stack.append(new_item)
                imported_count += 1

            if imported_count > 0:
                self.save_data()
                QMessageBox.information(self, "成功", f"成功合并导入 {imported_count} 条规则到 {tree_name}。")
            else:
                QMessageBox.warning(self, "导入提示", "在文件中没有找到有效的数据行。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入过程中发生错误: {e}")

    def export_data(self):
        if not openpyxl:
            QMessageBox.critical(self, "缺少库", "导出功能需要 'openpyxl' 库。\n请通过命令 'pip install openpyxl' 安装。")
            return

        items = [f"普通({self.term})奖励", f"特殊({self.term})奖励"]
        item, ok = QInputDialog.getItem(self, "选择导出", "您想从哪个列表导出规则？", items, 0, False)
        if not ok or not item:
            return

        tree = self.normal_rules_tree if item == f"普通({self.term})奖励" else self.special_rules_tree
        default_name = item

        file_path, _ = QFileDialog.getSaveFileName(self, f"导出 {default_name} 到 Excel", f"{default_name}.xlsx",
                                                   "Excel 文件 (*.xlsx)")
        if not file_path:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = default_name

            sheet["A1"] = "规则内容"
            sheet["B1"] = "对应奖励"
            sheet["C1"] = f"对应{self.term}"
            sheet.column_dimensions[get_column_letter(1)].width = 50
            sheet.column_dimensions[get_column_letter(2)].width = 30
            sheet.column_dimensions[get_column_letter(3)].width = 15

            def write_items_recursive(parent_item, current_row, level=0):
                for i in range(parent_item.childCount()):
                    item = parent_item.child(i)
                    indent = "  " * level
                    plan_text = f"{indent}{item.text(0)}"
                    reward_text = item.text(1)
                    reward_val_str = item.text(2)

                    sheet[f"A{current_row}"] = plan_text
                    sheet[f"B{current_row}"] = reward_text
                    try:
                        sheet[f"C{current_row}"] = int(reward_val_str)
                    except (ValueError, TypeError):
                        sheet[f"C{current_row}"] = reward_val_str

                    current_row += 1
                    current_row = write_items_recursive(item, current_row, level + 1)
                return current_row

            write_items_recursive(tree.invisibleRootItem(), 2)
            workbook.save(file_path)
            QMessageBox.information(self, "成功", f"{default_name} 已成功导出到:\n{file_path}")

        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出过程中发生错误: {e}")

